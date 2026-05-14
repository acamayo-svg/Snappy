# -*- coding: utf-8 -*-
"""Genera casos_prueba_snappy.xlsx (estilo hoja de calculo)."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

OUT = "casos_prueba_snappy.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FONT = Font(bold=True, size=14)


def style_header_row(ws, row=1, max_col=None):
    max_col = max_col or ws.max_column
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_grid(ws, start_row=1, end_row=None, start_col=1, end_col=None):
    end_row = end_row or ws.max_row
    end_col = end_col or ws.max_column
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = BORDER


def autosize_columns(ws, min_width=10, max_width=55):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = min_width
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            if v is None:
                continue
            max_len = min(max_width, max(max_len, len(str(v)) * 0.9 + 2))
        ws.column_dimensions[letter].width = max_len


CASOS = [
    {
        "id": "CP-EST-01-FELIZ",
        "nombre": "Registro exitoso de establecimiento",
        "tipo": "Feliz",
        "hu": "HU-EST-01",
        "desc": "Validar registro con datos validos, correo nuevo, codigo correcto y contrasena valida.",
        "pre": "Establecimiento no registrado. Correo operativo para codigo.",
        "rel": "HU-EST-01 - Registrar establecimiento",
        "estado": "Pendiente de ejecucion",
        "pasos": [
            ("Ingresar al formulario de registro de establecimiento.", "Se muestra el formulario de registro."),
            ("Diligenciar datos basicos validos y correo no registrado.", "Permite continuar a verificacion."),
            ("Solicitar envio de codigo de verificacion.", "Envia codigo al correo."),
            ("Ingresar codigo de verificacion correcto.", "Valida codigo y habilita contrasena."),
            ("Definir contrasena valida y confirmar registro.", "Crea la cuenta exitosamente."),
        ],
        "com": "Caso feliz HU-EST-01.",
    },
    {
        "id": "CP-EST-01-NOFELIZ",
        "nombre": "Registro fallido por correo ya registrado",
        "tipo": "No feliz",
        "hu": "HU-EST-01",
        "desc": "Impedir registro cuando el correo ya existe.",
        "pre": "Ya existe cuenta con el correo de prueba.",
        "rel": "HU-EST-01 - Registrar establecimiento",
        "estado": "Pendiente de ejecucion",
        "pasos": [
            ("Ingresar al formulario de registro.", "Se muestra el formulario."),
            ("Diligenciar datos basicos validos.", "Acepta datos basicos."),
            ("Ingresar correo ya registrado y continuar.", 'Mensaje "Correo ya existe".'),
            ("Intentar avanzar a verificacion.", "No permite continuar."),
            ("Intentar finalizar registro.", "No crea la cuenta."),
        ],
        "com": 'No feliz: correo duplicado.',
    },
    {
        "id": "CP-EST-02-FELIZ",
        "nombre": "Registro exitoso de producto en catalogo",
        "tipo": "Feliz",
        "hu": "HU-EST-02",
        "desc": "Registrar producto valido; queda Activo y asociado al establecimiento.",
        "pre": "Establecimiento autenticado en gestion de menu.",
        "rel": "HU-EST-02 - Registrar producto en el catalogo",
        "estado": "Pendiente de ejecucion",
        "pasos": [
            ("Ingresar al panel de gestion de menu.", "Opcion para agregar producto."),
            ('Seleccionar "Agregar producto".', "Formulario de producto."),
            ("Ingresar nombre, precio > 0, unidades, categoria, descripcion.", "Valida campos."),
            ("Confirmar guardado.", "Registra el producto."),
            ("Consultar en catalogo.", "Aparece Activo y del establecimiento."),
        ],
        "com": "Caso feliz HU-EST-02.",
    },
    {
        "id": "CP-EST-02-NOFELIZ",
        "nombre": "Registro fallido por precio invalido",
        "tipo": "No feliz",
        "hu": "HU-EST-02",
        "desc": "Rechazar precio igual o menor a cero.",
        "pre": "Establecimiento autenticado.",
        "rel": "HU-EST-02 - Registrar producto en el catalogo",
        "estado": "Pendiente de ejecucion",
        "pasos": [
            ('Abrir "Agregar producto".', "Formulario visible."),
            ("Diligenciar nombre, unidades, categoria, descripcion validos.", "Acepta campos."),
            ("Ingresar precio 0 o negativo.", "Detecta precio invalido."),
            ("Confirmar guardado.", "Mensaje de error de validacion."),
            ("Ver listado de productos.", "No se registra el producto."),
        ],
        "com": "No feliz: precio invalido.",
    },
    {
        "id": "CP-PED-04-FELIZ",
        "nombre": "Visualizacion correcta de pedidos recibidos",
        "tipo": "Feliz",
        "hu": "HU-CLI-04",
        "desc": "Ver pedidos con datos completos por pestanas.",
        "pre": "Establecimiento autenticado con pedidos existentes.",
        "rel": "HU-CLI-04 - Visualizar pedidos recibidos",
        "estado": "Pendiente de ejecucion",
        "pasos": [
            ("Ingresar a pedidos recibidos.", "Pestanas PAGADOS, LISTOS, EN CAMINO, ENTREGADOS."),
            ("Seleccionar cada pestana.", "Lista pedidos del estado."),
            ("Abrir detalle de un pedido.", "Numero, productos, monto, estado, direccion, telefono."),
            ('Usar boton "Actualizar".', "Refresca la vista."),
            ("Validar filtro por establecimiento.", "Solo pedidos del establecimiento autenticado."),
        ],
        "com": "Caso feliz HU-CLI-04 (establecimiento).",
    },
    {
        "id": "CP-PED-04-NOFELIZ",
        "nombre": "Datos incompletos del cliente (No especificado)",
        "tipo": "No feliz",
        "hu": "HU-CLI-04",
        "desc": 'Si falta direccion o telefono, mostrar "No especificado".',
        "pre": "Pedido donde falte direccion o telefono del cliente.",
        "rel": "HU-CLI-04 - Visualizar pedidos recibidos",
        "estado": "Pendiente de ejecucion",
        "pasos": [
            ("Ingresar a pedidos recibidos.", "Carga pedidos."),
            ("Abrir pedido con datos incompletos.", "Muestra detalle."),
            ("Revisar direccion.", 'Si falta: "No especificado".'),
            ("Revisar telefono.", 'Si falta: "No especificado".'),
            ("Revisar resto del pedido.", "Numero, productos, monto, estado visibles sin error."),
        ],
        "com": "No feliz: datos cliente incompletos.",
    },
    {
        "id": "CP-CLI-06-FELIZ",
        "nombre": "Calificar y guardar reseña con puntuación válida",
        "tipo": "Feliz",
        "hu": "HU-CLI-06",
        "desc": "Cliente califica pedido Entregado con puntuación 1-5; comentario opcional; sistema guarda y confirma.",
        "pre": "Cliente autenticado; pedido en estado Entregado; opción Calificar y dejar reseña habilitada.",
        "rel": "HU-CLI-06 - Calificar y dejar reseña del pedido",
        "estado": "Pendiente de ejecucion",
        "pasos": [
            ("Acceder a la opción Calificar y dejar reseña del pedido entregado.", "Se muestra formulario de calificación y comentario."),
            ("Ingresar puntuación entre 1 y 5 (ej. 4).", "El sistema acepta el valor."),
            ("Opcional: ingresar comentario o dejarlo vacío.", "Permite continuar (comentario opcional)."),
            ("Confirmar envío de la reseña.", "El sistema guarda la reseña asociada al pedido."),
            ("Verificar mensaje o confirmación en pantalla.", "El sistema confirma que la reseña fue guardada exitosamente."),
        ],
        "com": "HU-CLI-06 feliz: pedido Entregado, puntuacion 1-5, comentario opcional, persistencia y confirmacion al usuario.",
    },
    {
        "id": "CP-CLI-06-NOFELIZ",
        "nombre": "Rechazo por puntuación fuera del rango 1 a 5",
        "tipo": "No feliz",
        "hu": "HU-CLI-06",
        "desc": "No guardar reseña si la puntuación está fuera de 1-5; no permitir continuar.",
        "pre": "Cliente autenticado; acceso a calificar (pedido Entregado según HU).",
        "rel": "HU-CLI-06 - Calificar y dejar reseña del pedido",
        "estado": "Pendiente de ejecucion",
        "pasos": [
            ("Acceder a Calificar y dejar reseña.", "Se muestra el formulario."),
            ("Ingresar puntuación fuera de rango (ej. 0, 6 o valor no permitido).", "El sistema detecta valor inválido."),
            ("Intentar confirmar o guardar la reseña.", "El sistema no guarda la reseña."),
            ("Intentar avanzar o cerrar sin corregir.", "No permite continuar con el flujo de guardado exitoso."),
            ("Verificar que el pedido no tenga reseña nueva registrada.", "No se persiste calificación inválida."),
        ],
        "com": "HU-CLI-06 no feliz: puntuacion fuera de 1-5; sin guardado en BD; UI bloquea flujo exitoso hasta corregir.",
    },
    {
        "id": "CP-EST-03-FELIZ",
        "nombre": "Actualizar precio de producto propio con valor válido",
        "tipo": "Feliz",
        "hu": "HU-EST-03",
        "desc": "Establecimiento autenticado actualiza precio de su producto (>0, no vacío); cambio visible y confirmación.",
        "pre": "Establecimiento autenticado; producto registrado en su catálogo.",
        "rel": "HU-EST-03 - Actualizar precio del producto",
        "estado": "Pendiente de ejecucion",
        "pasos": [
            ("Ingresar al panel de gestión del menú / catálogo.", "Lista de productos del establecimiento."),
            ("Seleccionar un producto propio y opción de editar o actualizar precio.", "Se muestra campo de precio con valor actual."),
            ("Ingresar nuevo precio mayor que cero (campo no vacío).", "El sistema valida el dato."),
            ("Confirmar guardado de la actualización.", "El sistema confirma actualización exitosa del precio."),
            ("Verificar catálogo visible para clientes (o vista pública).", "El cambio se refleja con el nuevo precio."),
        ],
        "com": "HU-EST-03 feliz: producto propio, precio obligatorio y >0, confirmacion, reflejo inmediato en catalogo cliente.",
    },
    {
        "id": "CP-EST-03-NOFELIZ",
        "nombre": "Rechazo al actualizar precio inválido o vacío",
        "tipo": "No feliz",
        "hu": "HU-EST-03",
        "desc": "No permitir precio vacío, cero o negativo; no actualizar hasta corregir.",
        "pre": "Establecimiento autenticado con producto propio en catálogo.",
        "rel": "HU-EST-03 - Actualizar precio del producto",
        "estado": "Pendiente de ejecucion",
        "pasos": [
            ("Abrir actualización de precio de un producto propio.", "Formulario o campo de precio visible."),
            ("Borrar el precio o ingresar cero / valor negativo.", "El sistema marca dato inválido u obligatorio."),
            ("Intentar guardar la actualización.", "No confirma actualización exitosa; muestra error de validación."),
            ("Verificar precio en catálogo del establecimiento.", "El precio anterior se mantiene (no aplica cambio inválido)."),
            ("Corregir a precio válido y guardar.", "Solo entonces aplica caso feliz; aquí se valida el rechazo previo."),
        ],
        "com": "HU-EST-03 no feliz: precio vacio, cero o negativo; error de validacion; precio publicado sin cambio.",
    },
]


def sheet_casos_completos(wb):
    ws = wb.create_sheet("Casos_prueba_completos", 0)
    ws.merge_cells("A1:S1")
    ws["A1"] = "Snappy - Casos de prueba (vista tipo Excel)"
    ws["A1"].font = TITLE_FONT

    headers = [
        "ID caso",
        "Nombre caso",
        "Tipo",
        "HU",
        "Descripcion",
        "Precondiciones",
        "Relaciones HU",
        "Estado CP",
        "Paso 1",
        "Resultado esperado 1",
        "Paso 2",
        "Resultado esperado 2",
        "Paso 3",
        "Resultado esperado 3",
        "Paso 4",
        "Resultado esperado 4",
        "Paso 5",
        "Resultado esperado 5",
        "Comentarios",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, row=3, max_col=len(headers))

    row = 4
    for cp in CASOS:
        pasos = cp["pasos"] + [("", "")] * (5 - len(cp["pasos"]))
        row_data = [
            cp["id"],
            cp["nombre"],
            cp["tipo"],
            cp["hu"],
            cp["desc"],
            cp["pre"],
            cp["rel"],
            cp["estado"],
        ]
        for i in range(5):
            row_data.extend([pasos[i][0], pasos[i][1]])
        row_data.append(cp["com"])
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    apply_grid(ws, start_row=3, end_row=row - 1, start_col=1, end_col=len(headers))
    ws.freeze_panes = "A4"
    autosize_columns(ws)


def sheet_decision_est01(wb):
    ws = wb.create_sheet("Decision_HU-EST-01")
    ws["A1"] = "Tabla de decision - HU-EST-01 Registrar establecimiento"
    ws["A1"].font = TITLE_FONT
    data = [
        ["Regla", "R1 Exito", "R2 Correo duplicado", "R3 Codigo incorrecto", "R4 Datos basicos invalidos"],
        ["C1: Datos basicos validos", "V", "V", "V", "F"],
        ["C2: Correo no registrado", "V", "F", "V", "V"],
        ["C3: Codigo verificacion correcto", "V", "N/A", "F", "N/A"],
        ["C4: Contrasena valida", "V", "N/A", "N/A", "N/A"],
        ["A1: Crea cuenta exitosamente", "X", "", "", ""],
        ["A2: Bloquea / mensaje error", "", "X", "X", "X"],
    ]
    start = 3
    for r, row in enumerate(data, start=start):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if r == start:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            cell.border = BORDER
    autosize_columns(ws)


def sheet_decision_est02(wb):
    ws = wb.create_sheet("Decision_HU-EST-02")
    ws["A1"] = "Tabla de decision - HU-EST-02 Registrar producto"
    ws["A1"].font = TITLE_FONT
    data = [
        ["Regla", "R1 Exito", "R2 Precio invalido", "R3 Nombre vacio"],
        ["C1: Establecimiento autenticado", "V", "V", "V"],
        ["C2: Nombre informado", "V", "V", "F"],
        ["C3: Precio > 0", "V", "F", "V"],
        ["C4: Unidades >= 0 entero", "V", "V", "V"],
        ["C5: Categoria valida", "V", "V", "V"],
        ["A1: Registra producto Activo", "X", "", ""],
        ["A2: No registra / error", "", "X", "X"],
    ]
    start = 3
    for r, row in enumerate(data, start=start):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if r == start:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            cell.border = BORDER
    autosize_columns(ws)


def sheet_decision_cli04(wb):
    ws = wb.create_sheet("Decision_HU-CLI-04")
    ws["A1"] = "Tabla de decision - HU-CLI-04 Visualizar pedidos"
    ws["A1"].font = TITLE_FONT
    data = [
        ["Regla", "R1 Datos completos", "R2 Falta direccion", "R3 Falta telefono", "R4 Faltan ambos"],
        ["C1: Establecimiento autenticado", "V", "V", "V", "V"],
        ["C2: Hay pedidos del establecimiento", "V", "V", "V", "V"],
        ["C3: Direccion registrada", "V", "F", "V", "F"],
        ["C4: Telefono registrado", "V", "V", "F", "F"],
        ["A1: Muestra datos completos en vista", "X", "", "", ""],
        ['A2: Muestra "No especificado" donde falte', "", "X", "X", "X"],
    ]
    start = 3
    for r, row in enumerate(data, start=start):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if r == start:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            cell.border = BORDER
    autosize_columns(ws)


def sheet_decision_cli06(wb):
    ws = wb.create_sheet("Decision_HU-CLI-06")
    ws["A1"] = "Tabla de decision - HU-CLI-06 Calificar y dejar reseña del pedido"
    ws["A1"].font = TITLE_FONT
    data = [
        ["Regla", "R1 Exito", "R2 Puntuacion fuera de rango", "R3 Pedido no entregado"],
        ["C1: Cliente autenticado", "V", "V", "V"],
        ["C2: Pedido en estado Entregado", "V", "V", "F"],
        ["C3: Puntuacion entre 1 y 5", "V", "F", "N/A"],
        ["C4: Comentario (opcional) valido si existe", "V", "V", "N/A"],
        ["A1: Guarda reseña y confirma exito", "X", "", ""],
        ["A2: No guarda / bloquea continuar", "", "X", "X"],
    ]
    start = 3
    for r, row in enumerate(data, start=start):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if r == start:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            cell.border = BORDER
    autosize_columns(ws)


def sheet_decision_est03(wb):
    ws = wb.create_sheet("Decision_HU-EST-03")
    ws["A1"] = "Tabla de decision - HU-EST-03 Actualizar precio del producto"
    ws["A1"].font = TITLE_FONT
    data = [
        ["Regla", "R1 Exito", "R2 Precio invalido o vacio", "R3 Producto no propio"],
        ["C1: Establecimiento autenticado", "V", "V", "V"],
        ["C2: Producto del catalogo del establecimiento (propio)", "V", "V", "F"],
        ["C3: Campo precio no vacio", "V", "F", "V"],
        ["C4: Precio mayor que cero", "V", "F", "V"],
        ["A1: Actualiza precio y confirma; refleja en catalogo", "X", "", ""],
        ["A2: No actualiza / error o sin permiso", "", "X", "X"],
    ]
    start = 3
    for r, row in enumerate(data, start=start):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if r == start:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            cell.border = BORDER
    autosize_columns(ws)


def main():
    wb = Workbook()
    # quitar hoja por defecto
    default = wb.active
    wb.remove(default)

    sheet_casos_completos(wb)
    sheet_decision_est01(wb)
    sheet_decision_est02(wb)
    sheet_decision_cli04(wb)
    sheet_decision_cli06(wb)
    sheet_decision_est03(wb)

    wb.save(OUT)
    print("Escrito:", OUT)


if __name__ == "__main__":
    main()
